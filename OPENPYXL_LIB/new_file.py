import pathlib
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

ROOT_FOLDER = pathlib.Path(__file__).parent
FILE_PATH = ROOT_FOLDER / 'my_file_excel.xlsx'

# Datas
students = [
    # nome      idade nota
    ['João',    14,   5.5],
    ['Maria',   13,   9.7],
    ['Luiz',    15,   8.8],
    ['Alberto', 16,   10],
]

# Workbook and worksheet.
workbook = Workbook()
worksheet: Worksheet = workbook.active #type: ignore

# Creating the headers of my sheet
worksheet.cell(1, 1, 'Name')
worksheet.cell(1, 2, 'age')
worksheet.cell(1, 3, 'note')

# Iterating over the data

# for i, row in enumerate(students, start=2):
#     for j, collumn in enumerate(row, 1):
#         worksheet.cell(i, j, collumn)

# Add data in my worksheet

for student in students:
    worksheet.append(student)


# Saving the file.
workbook.save(FILE_PATH) 


print(FILE_PATH)